from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT = "Malgun Gothic"
BLUE = "1F4E79"
GREEN, YELLOW, RED, GREY = "D5E8D4", "FFF2CC", "F8CECC", "F2F2F2"
thin = Side(style="thin", color="BFBFBF")
B = Border(left=thin, right=thin, top=thin, bottom=thin)
hfill = PatternFill("solid", fgColor=BLUE)
hfont = Font(name=FONT, bold=True, color="FFFFFF", size=10)
cfont = Font(name=FONT, size=10)
wrap = Alignment(vertical="top", wrap_text=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

gF = PatternFill("solid", fgColor=GREEN); yF = PatternFill("solid", fgColor=YELLOW)
rF = PatternFill("solid", fgColor=RED); kF = PatternFill("solid", fgColor=GREY)

wb = Workbook()

def style_header(ws, ncol, row=1):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = hfill; cell.font = hfont; cell.border = B
        cell.alignment = center
    ws.row_dimensions[row].height = 30

def fill_rows(ws, data, startrow=2):
    for r, rowvals in enumerate(data, start=startrow):
        for c, v in enumerate(rowvals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = cfont; cell.border = B; cell.alignment = wrap

def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def rag(ws, col_letter, nrows, green=(), yellow=(), red=(), grey=()):
    rng = f"{col_letter}2:{col_letter}{nrows+1}"
    for vals, fill in [(green, gF), (yellow, yF), (red, rF), (grey, kF)]:
        for v in vals:
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{v}"'], fill=fill))

def dropdown(ws, col_letter, nrows, options):
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{nrows+1}")

def finalize(ws, ncol, nrows, freeze="A2"):
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{nrows+1}"
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False

# ---------- Index ----------
ix = wb.active; ix.title = "Index"
ix.sheet_view.showGridLines = False
ix["B2"] = "KT \u00D7 VNPT/MyTV \u2014 PMO Trackers"
ix["B2"].font = Font(name=FONT, bold=True, size=16, color=BLUE)
ix["B3"] = "AI-Powered MyTV \u2014 KT \u00D7 VNPT Joint Platform Initiative  |  CONFIDENTIAL \u2014 KT Global Business Division"
ix["B3"].font = Font(name=FONT, italic=True, size=9, color="808080")
rows = [
    ("Tab", "Purpose"),
    ("RTM", "Requirements Traceability Matrix \u2014 VNPT RFI 7 items \u2192 KT response, evidence, owner, status"),
    ("Action_Tracker", "Action items (incl. session draft AI-01\u201309) \u2014 owner, target, status, decision link"),
    ("Decision_Log", "Cumulative decisions \u2192 linked contract clause (Contract Traceability)"),
    ("RAID", "Risks / Assumptions / Issues / Dependencies \u2014 with grade & response"),
    ("Deliverable_Registry", "All project deliverables \u2014 author, approver, version, status (registration)"),
]
for r, (a, b) in enumerate(rows, start=5):
    ca, cb = ix.cell(row=r, column=2, value=a), ix.cell(row=r, column=3, value=b)
    ca.border = B; cb.border = B; ca.alignment = wrap; cb.alignment = wrap
    if r == 5:
        ca.fill = hfill; ca.font = hfont; cb.fill = hfill; cb.font = hfont
    else:
        ca.font = Font(name=FONT, bold=True, size=10); cb.font = cfont
ix.cell(row=12, column=2, value="RAG legend").font = Font(name=FONT, bold=True, size=10)
leg = [("Green", "Done / Confirmed / Approved / Closed / Low", gF),
       ("Yellow", "In progress / Tentative / In review / Mitigating / Med / Not started", yF),
       ("Red", "Blocked / Delayed / High", rF)]
for r, (lab, desc, f) in enumerate(leg, start=13):
    c1 = ix.cell(row=r, column=2, value=lab); c1.fill = f; c1.border = B; c1.font = cfont
    c2 = ix.cell(row=r, column=3, value=desc); c2.font = cfont; c2.border = B; c2.alignment = wrap
ix.cell(row=17, column=2, value="Note").font = Font(name=FONT, bold=True, size=10)
ix.cell(row=17, column=3, value="Seed rows reflect the WS1 working-session package; update as the project progresses. Cost/effort = planning estimates, final figures via separate quotation. External facts (regulation, certification, benchmarks) require source + Tier per the PMO citation rule.").font = cfont
ix.cell(row=17, column=3).alignment = wrap
widths(ix, [3, 22, 95])

# ---------- RTM ----------
ws = wb.create_sheet("RTM")
hdr = ["ID", "RFI#", "Requirement (summary)", "Mgmt Element", "KT response / approach", "Evidence doc", "Owner", "Status", "Notes"]
ws.append(hdr)
rtm = [
 ["RTM-01", 1, "Product portfolio, specs, target, costing", "1,3", "STB 4 / Soundbar spec sheets + cost framework (planning est.)", "STB Spec; Cost & BM model", "KT", "In progress", ""],
 ["RTM-02", 2, "Media AI Agent integration scope & model, OTA", "2,1", "Cloud-based APK + legacy-HW constraints noted; OTA via TR-369", "AI Agent Spec; IDD", "KT", "In progress", "8/15 features cloud-deployable; 4 blocked (NPU)"],
 ["RTM-03", 3, "OS/launcher, Google Cert roadmap, dev R&R", "2,4,10,5", "GTV4O vs Custom comparison + cert roadmap + R&R", "OS/Launcher strategy; Cert roadmap", "Joint", "Not started", "Launcher dev R&R to confirm"],
 ["RTM-04", 4, "BM cost breakdown, MG, MOQ, Revenue Share", "3,12", "OEM\u2192License hybrid + pricing/MG + commercial terms", "BM comparison; Contract annex", "Joint", "In progress", "VNPT priority \u2014 device pricing, MG, terms"],
 ["RTM-05", 5, "TMS functions, deployment, OTA targeting", "1,2", "TR-069/369; on-prem/cloud; segment/region OTA", "TMS Spec; IDD", "KT", "Not started", ""],
 ["RTM-06", 6, "Logistics, MIC ICT cert, AS, AFR, Epidemic", "8,10,7", "Customs/warehousing + MIC + AFR target + Epidemic policy", "Logistics plan; SLA def.", "Joint", "Not started", ""],
 ["RTM-07", 7, "Data sovereignty \u2014 VNPT Cloud, Decree 13/2023", "10,2", "Full deployment on VNPT Cloud (VN); data-residency design", "Data sovereignty design", "Joint", "Not started", "Mandatory requirement (hard filter)"],
]
fill_rows(ws, rtm)
widths(ws, [9, 6, 30, 12, 34, 22, 9, 13, 30])
rag(ws, "H", len(rtm), green=["Done"], yellow=["In progress", "Not started"], red=["Blocked"])
dropdown(ws, "H", len(rtm), ["Not started", "In progress", "Done", "Blocked"])
dropdown(ws, "G", len(rtm), ["KT", "MyTV", "VNPT", "VNPT/MyTV", "Joint", "PMO"])
finalize(ws, len(hdr), len(rtm))

# ---------- Action_Tracker ----------
ws = wb.create_sheet("Action_Tracker")
hdr = ["ID", "Action item", "Area", "Owner", "Target", "Status", "Decision link", "Notes"]
ws.append(hdr)
act = [
 ["AI-01", "Issue STB pricing + MG proposal (planning estimate)", "BM", "KT", "Pre / at session", "Open", "DEC-01", ""],
 ["AI-02", "Provide target volumes & acceptable MG appetite", "BM", "VNPT/MyTV", "Pre / at session", "Open", "DEC-01", ""],
 ["AI-03", "Confirm OS/launcher direction (GTV4O vs Custom) & dev R&R", "STB", "Joint", "Session", "Open", "DEC-03", ""],
 ["AI-04", "Share legacy-system docs for integration analysis", "AI/TMS", "VNPT/MyTV", "+1 week", "Open", "", ""],
 ["AI-05", "Deliver Media AI Agent feature scope incl. legacy-HW feasibility", "AI/TMS", "KT", "Session / +1wk", "Open", "", ""],
 ["AI-06", "Define TMS deployment (on-prem/cloud) & OTA targeting requirements", "AI/TMS", "Joint", "+1 week", "Open", "", ""],
 ["AI-07", "Align certification (MIC/Google) & data-sovereignty (VNPT Cloud) ownership", "Cert/Data", "Joint", "+2 weeks", "Open", "DEC-02", ""],
 ["AI-08", "Execute mutual NDA (VNPT standard terms)", "Legal", "Both \u00B7 Legal", "Before deep-dive", "Open", "DEC-05", "VNPT to send standard NDA"],
 ["AI-09", "Schedule Requirements Workshop (Stage 4)", "Plan", "Joint", "Session", "Open", "DEC-04", "Teams invite by VNPT"],
]
fill_rows(ws, act)
widths(ws, [8, 46, 10, 13, 16, 12, 13, 26])
rag(ws, "F", len(act), green=["Done"], yellow=["In progress", "Open"], red=["Blocked", "Overdue"])
dropdown(ws, "F", len(act), ["Open", "In progress", "Done", "Blocked", "Overdue"])
dropdown(ws, "D", len(act), ["KT", "MyTV", "VNPT", "VNPT/MyTV", "Joint", "PMO", "Both \u00B7 Legal"])
finalize(ws, len(hdr), len(act))

# ---------- Decision_Log ----------
ws = wb.create_sheet("Decision_Log")
hdr = ["DEC-ID", "Date", "Meeting", "Stage", "Decision", "Rationale", "Owner", "Status", "Linked contract clause", "Linked deliverable"]
ws.append(hdr)
dec = [
 ["DEC-01", "", "Working Session", 3, "STB business model direction = OEM\u2192License hybrid (pricing/MG/terms to finalize)", "CAPEX/risk minimization; phased commitment", "Joint", "Tentative", "\u00A7Pricing & BM, License, MG", "BM comparison"],
 ["DEC-02", "", "Working Session", 3, "User/behavioral data stored & processed on VNPT Cloud (Vietnam)", "Decree 13/2023 data residency (mandatory)", "Joint", "Tentative", "\u00A7Data sovereignty & SLA", "Data sovereignty design"],
 ["DEC-03", "", "Working Session", 3, "Launcher: GTV4O prioritized; dev R&R to be confirmed", "Google policy; UX customization vs effort", "Joint", "Tentative", "\u00A7Scope & R&R", "OS/Launcher strategy"],
 ["DEC-04", "", "Working Session", 3, "Proceed to Requirements Workshop (Stage 4); MOU target Q3 2026", "Move from deep-dive to requirements baseline", "Joint", "Tentative", "\u00A7Term & milestones", "Requirements analysis"],
 ["DEC-05", "", "Working Session", 3, "Mutual NDA executed on VNPT standard terms before deep-dive exchange", "Manage legal risk on confidential exchange", "Both \u00B7 Legal", "Tentative", "NDA (standalone)", "Mutual NDA"],
]
fill_rows(ws, dec)
widths(ws, [9, 11, 16, 7, 40, 28, 13, 12, 24, 22])
rag(ws, "H", len(dec), green=["Confirmed"], yellow=["Tentative"], red=["Rejected"])
dropdown(ws, "H", len(dec), ["Tentative", "Confirmed", "Rejected", "Superseded"])
finalize(ws, len(hdr), len(dec))

# ---------- RAID ----------
ws = wb.create_sheet("RAID")
hdr = ["ID", "Type", "Description", "Impact", "Likelihood", "Grade", "Response", "Owner", "Due", "Status"]
ws.append(hdr)
raid = [
 ["RA-01", "Risk", "MIC ICT / Google certification delay", "High", "Med", "High", "Parallel tracks; early start; schedule buffer", "Joint", "P2", "Open"],
 ["RA-02", "Risk", "Data sovereignty not met (Decree 13/2023)", "High", "Med", "High", "VNPT Cloud-first design; legal review", "Joint", "P2", "Open"],
 ["RA-03", "Risk", "Legacy STB HW (NPU absent) blocks some AI features", "Med", "High", "Med", "Cloud APK alternative; agree feature scope", "KT", "P1", "Open"],
 ["RA-04", "Issue", "Launcher development R&R undecided (RFI #3)", "Med", "High", "Med", "RACI agreement at/after session (AI-03)", "PMO", "Session", "Open"],
 ["RA-05", "Risk", "MG / commercial terms alignment gap", "Med", "Med", "Med", "Early pricing+MG proposal; volume alignment", "Joint", "P0", "Open"],
 ["RA-06", "Depend.", "WS2\u20136 depend on WS1 milestones (hub)", "Med", "High", "Med", "Gate at MS-5 (UAT); track dependency", "PMO", "Ongoing", "Tracking"],
]
fill_rows(ws, raid)
widths(ws, [8, 9, 38, 9, 11, 8, 34, 9, 11, 11])
rag(ws, "F", len(raid), green=["Low"], yellow=["Med"], red=["High"])
rag(ws, "J", len(raid), green=["Closed"], yellow=["Open", "Mitigating", "Tracking"], red=["Escalated"])
dropdown(ws, "B", len(raid), ["Risk", "Assumption", "Issue", "Depend."])
dropdown(ws, "F", len(raid), ["Low", "Med", "High"])
dropdown(ws, "J", len(raid), ["Open", "Mitigating", "Tracking", "Closed", "Escalated"])
finalize(ws, len(hdr), len(raid))

# ---------- Deliverable_Registry ----------
ws = wb.create_sheet("Deliverable_Registry")
hdr = ["DEL-ID", "Deliverable", "Stage", "Phase", "Author (R)", "Approver (A)", "Version", "Status", "Reg. date", "Linked (DEC/meeting)"]
ws.append(hdr)
TODAY = "2026-06-06"
deli = [
 ["DEL-01", "WS1 Joint Working Session Package", 3, "P0", "KT (PMO)", "KT", "v1.0", "Approved", TODAY, "Working Session"],
 ["DEL-02", "RFI 7-item response (RTM)", 3, "P0", "KT", "Joint", "v0.3", "Drafting", TODAY, "RTM"],
 ["DEL-03", "STB pricing & MG model", 3, "P0", "KT", "KT", "v0.1", "Drafting", "", "DEC-01"],
 ["DEL-04", "STB deployment plan", 3, "P0\u2192P1", "KT", "Joint", "v0.1", "Drafting", "", "AI-03"],
 ["DEL-05", "Media AI Agent & TMS scope/impl. plan", 3, "P0\u2192P1", "KT", "Joint", "v0.1", "Drafting", "", "AI-05/06"],
 ["DEL-06", "Interface Definition Document (IDD)", 7, "P1", "KT", "Joint", "\u2014", "Drafting", "", ""],
 ["DEL-07", "Certification & regulatory roadmap", 10, "P2", "Joint", "Joint", "\u2014", "Drafting", "", "DEC-02"],
 ["DEL-08", "Mutual NDA (VNPT standard)", 3, "P0", "VNPT/MyTV", "Both \u00B7 Legal", "\u2014", "In review", "", "DEC-05"],
]
fill_rows(ws, deli)
widths(ws, [9, 38, 7, 8, 12, 13, 9, 12, 11, 18])
rag(ws, "H", len(deli), green=["Approved"], yellow=["Drafting", "In review"], red=["Rejected"])
dropdown(ws, "H", len(deli), ["Drafting", "In review", "Approved", "Rejected"])
finalize(ws, len(hdr), len(deli))

wb.save("/home/claude/pack/KT_VNPT_WS1_PMO_Trackers.xlsx")
print("XLSX BUILT")
