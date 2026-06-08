from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT="Malgun Gothic"; NAVY="1F4E79"
GREEN,YELLOW,RED,GREY,AMBER,LTBLUE="D5E8D4","FFF2CC","F8CECC","EDEDED","FCE4D6","DDEBF7"
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
hfill=PatternFill("solid",fgColor=NAVY); hfont=Font(name=FONT,bold=True,color="FFFFFF",size=10)
cfont=Font(name=FONT,size=10); wrap=Alignment(vertical="top",wrap_text=True)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
gF,yF,rF,kF,aF,bF=(PatternFill("solid",fgColor=x) for x in (GREEN,YELLOW,RED,GREY,AMBER,LTBLUE))
MONEY='#,##0.00'; PCT='0.0%'

wb=Workbook()

def hdr(ws,headers,row=1):
    ws.append(headers) if row==1 else None
    for c in range(1,len(headers)+1):
        cell=ws.cell(row=row,column=c); cell.fill=hfill; cell.font=hfont; cell.border=B; cell.alignment=ctr
    ws.row_dimensions[row].height=32
def widths(ws,ws_w):
    for i,w in enumerate(ws_w,1): ws.column_dimensions[get_column_letter(i)].width=w
def dd(ws,col,n,opts,start=2):
    dv=DataValidation(type="list",formula1='"'+",".join(opts)+'"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"{col}{start}:{col}{n}")

# ============ GUIDE ============
g=wb.active; g.title="Guide"; g.sheet_view.showGridLines=False
g["B2"]="KT × VNPT — Quotation & Settlement Reconciliation"; g["B2"].font=Font(name=FONT,bold=True,size=15,color=NAVY)
g["B3"]="협력사 → KT → VNPT 정산 무결성 워크북  |  CONFIDENTIAL"; g["B3"].font=Font(name=FONT,italic=True,size=9,color="808080")
rows=[
 ("",""),
 ("목적","견적·세금·계약·송장·지급 간 불일치(정산 오류)를 사전 차단한다. 견적서가 '단일 원가 기준'이며 정산 대사 탭이 이를 N-way로 검증한다."),
 ("",""),
 ("탭 구성","① Quotation(견적서·단일 원가 기준)  ② Tax_Cascade(세금 캐스케이드)  ③ Reconciliation(정산 대사 N-way)  ④ Back_to_Back(백투백 flow-down 체크)"),
 ("정산 체인","업스트림 협력사→KT(KRW/USD·한국 세제)  /  다운스트림 KT→VNPT(USD/VND·베트남 FCT). 백투백 원칙으로 KT 중간 리스크 최소화."),
 ("통화·환","라인별 통화 + USD 환산(FX rate) 입력 → 합계는 USD 기준. 환율 기준일을 통일한다."),
 ("핵심 통제","(1)견적 단일 원가 기준·scope 1:1  (2)세금 분류(goods/royalty/service) 체인 일관·gross-up·DTA  (3)백투백 flow-down  (4)견적↔계약↔검수↔송장↔지급 N-way 대사  (5)검수연동 지급"),
 ("",""),
 ("⚠ 주의","모든 금액·세율은 ILLUSTRATIVE(예시·계획 추정치)이며 정식 견적·세무 처리는 별도다. 세율·세무는 세무사, 계약 조항은 변호사(L5-bd-contract-review) 확인 전제."),
]
r=5
for k,v in rows:
    if k:
        a=g.cell(row=r,column=2,value=k); a.font=Font(name=FONT,bold=True,size=10,color=NAVY); a.alignment=wrap
        c=g.cell(row=r,column=3,value=v); c.font=cfont; c.alignment=wrap
    r+=1
widths(g,[2,14,110])

# ============ QUOTATION ============
q=wb.create_sheet("Quotation"); q.sheet_view.showGridLines=False
H=["Line ID","Solution Part","Item / 설명","Type","Unit","Qty","Ccy","FX→USD","협력사 단가\n(원ccy)","협력사 합계\n(USD)","KT 마진%","KT 마진\n(USD)","VNPT 가격\n(USD)","FCT 분류","Scope/SOW ref","Notes"]
hdr(q,H)
# example rows (ILLUSTRATIVE) — amber
ex=[
 ["Q-001","STB (Genie STB 4)","셋톱 단말 OEM","goods","ea",1000,"USD",1.0,100,None,0.10,None,None,"goods (1%+1%)","RTM-01","ILLUSTRATIVE — 교체"],
 ["Q-002","Media AI Agent","AI Agent SW License","royalty","license",1,"USD",1.0,50000,None,0.15,None,None,"royalty CIT 10%","RTM-02","ILLUSTRATIVE — 교체"],
 ["Q-003","SI / Integration","연동·구축 서비스","service","LS",1,"USD",1.0,80000,None,0.12,None,None,"service 5%+5%","RTM-05","ILLUSTRATIVE — 교체"],
]
for i,row in enumerate(ex,2): 
    for c,v in enumerate(row,1):
        cell=q.cell(row=i,column=c,value=v); cell.font=cfont; cell.border=B; cell.alignment=wrap; cell.fill=aF
# formulas for example + blank rows up to 21
last=21
for i in range(2,last+1):
    # J 협력사합계USD = I*F*H ; L set ; M 마진=J*L ; N=J+M
    q.cell(row=i,column=10,value=f"=IFERROR(I{i}*F{i}*H{i},\"\")")
    q.cell(row=i,column=12,value=f"=IFERROR(J{i}*K{i},\"\")")
    q.cell(row=i,column=13,value=f"=IFERROR(J{i}+L{i},\"\")")
    for c in range(1,len(H)+1):
        cell=q.cell(row=i,column=c)
        if cell.value is None: cell.font=cfont; cell.border=B; cell.alignment=wrap
        if c in (9,10,12,13): cell.number_format=MONEY
        if c==11: cell.number_format=PCT
# totals
q.cell(row=last+1,column=2,value="TOTAL (USD)").font=Font(name=FONT,bold=True,size=10)
for col in (10,12,13):
    L=get_column_letter(col); t=q.cell(row=last+1,column=col,value=f"=SUM({L}2:{L}{last})"); t.font=Font(name=FONT,bold=True); t.number_format=MONEY; t.fill=bF; t.border=B
q.cell(row=last+1,column=2).fill=bF; q.cell(row=last+1,column=2).border=B
widths(q,[8,16,22,9,6,7,6,8,13,13,8,13,13,13,12,16])
dd(q,"D",last,["goods","royalty","service"]); dd(q,"G",last,["KRW","USD","VND"])
q.freeze_panes="A2"; q.auto_filter.ref=f"A1:P{last}"

# ============ TAX_CASCADE ============
t=wb.create_sheet("Tax_Cascade"); t.sheet_view.showGridLines=False
t["A1"]="세율·세무 처리는 ILLUSTRATIVE 기본값(연구 기반)이며 적용은 사안별 — 반드시 세무사 확인. 베트남 상세: vietnam-legal-checklist.md §1"
t.merge_cells("A1:L1"); t["A1"].font=Font(name=FONT,bold=True,size=9,color="B00020"); t["A1"].alignment=wrap; t.row_dimensions[1].height=26
H2=["Line ID","Solution Part","Type","[한국] VAT%","[한국] 원천%\n(해외협력사)","이전가격\nTP","[VN] FCT VAT%","[VN] FCT CIT%","FCT 분류 근거","DTA 적용\nCIT%","Gross-up","세부담 주체","마진 영향 메모"]
hdr(t,H2,row=2)
tex=[
 ["Q-001","STB","goods",0.10,0,"N",0.01,0.01,"goods (기기)","-","계약 명시","VNPT 원천","HW 저율"],
 ["Q-002","AI Agent","royalty",0.10,0,"확인",0.0,0.10,"SW license=royalty","DTA 경감 검토","Y(권장)","협의","로열티 10% → DTA로 경감 가능"],
 ["Q-003","SI/Integration","service",0.10,0,"N",0.05,0.05,"service","-","협의","협의","서비스 5%+5%"],
]
for i,row in enumerate(tex,3):
    for c,v in enumerate(row,1):
        cell=t.cell(row=i,column=c,value=v); cell.font=cfont; cell.border=B; cell.alignment=wrap; cell.fill=aF
        if c in (4,5,7,8,10) and isinstance(v,(int,float)): cell.number_format=PCT
for i in range(6,16):
    for c in range(1,len(H2)+1):
        cell=t.cell(row=i,column=c); cell.font=cfont; cell.border=B; cell.alignment=wrap
        if c in (4,5,7,8,10): cell.number_format=PCT
dd(t,"C",15,["goods","royalty","service"],start=3); dd(t,"F",15,["Y","N","확인"],start=3); dd(t,"K",15,["Y","N","Y(권장)","협의"],start=3)
widths(t,[8,14,9,9,10,7,10,10,16,11,10,11,26])
t.freeze_panes="A3"

# ============ RECONCILIATION (N-way) ============
rc=wb.create_sheet("Reconciliation"); rc.sheet_view.showGridLines=False
rc["A1"]="N-way 대사: (1)견적 ↔ (2)계약/PO ↔ (3)검수 ↔ (4)송장 ↔ (5)지급 — 전부 일치해야 'Matched'. 검수연동 지급·변경통제 준수."
rc.merge_cells("A1:M1"); rc["A1"].font=Font(name=FONT,bold=True,size=9,color=NAVY); rc["A1"].alignment=wrap; rc.row_dimensions[1].height=26
H3=["Line/MS ID","Solution Part / 설명","Ccy","(1)견적","(2)계약/PO","(3)검수","(4)송장","(5)지급","Match","Variance","검수일","지급일","Notes"]
hdr(rc,H3,row=2)
rex=[
 ["Q-001","STB 단말 OEM","USD",100000,100000,100000,100000,0,None,None,"2026-09-01","","ILLUSTRATIVE"],
 ["Q-002","AI Agent License","USD",57500,57500,57500,0,0,None,None,"","","ILLUSTRATIVE"],
]
for i,row in enumerate(rex,3):
    for c,v in enumerate(row,1):
        cell=rc.cell(row=i,column=c,value=v); cell.font=cfont; cell.border=B; cell.alignment=wrap; cell.fill=aF
        if c in (4,5,6,7,8): cell.number_format=MONEY
last3=22
for i in range(3,last3+1):
    # Match: all of D..H equal -> Matched else CHECK (treat blanks)
    rc.cell(row=i,column=9,value=f'=IF(COUNT(D{i}:H{i})<5,"…",IF(AND(D{i}=E{i},E{i}=F{i},F{i}=G{i},G{i}=H{i}),"Matched","CHECK"))')
    rc.cell(row=i,column=10,value=f'=IFERROR(MAX(D{i}:H{i})-MIN(D{i}:H{i}),"")')
    for c in range(1,len(H3)+1):
        cell=rc.cell(row=i,column=c)
        if cell.value is None: cell.font=cfont; cell.border=B; cell.alignment=wrap
        if c in (4,5,6,7,8,10): cell.number_format=MONEY
        if c==9: cell.alignment=ctr
rc.cell(row=last3+1,column=2,value="TOTAL (USD)").font=Font(name=FONT,bold=True)
for col in (4,5,6,7,8):
    L=get_column_letter(col); cell=rc.cell(row=last3+1,column=col,value=f"=SUM({L}3:{L}{last3})"); cell.font=Font(name=FONT,bold=True); cell.number_format=MONEY; cell.fill=bF; cell.border=B
rc.cell(row=last3+1,column=2).fill=bF; rc.cell(row=last3+1,column=2).border=B
# conditional formatting on Match
rng=f"I3:I{last3}"
rc.conditional_formatting.add(rng,CellIsRule(operator="equal",formula=['"Matched"'],fill=gF))
rc.conditional_formatting.add(rng,CellIsRule(operator="equal",formula=['"CHECK"'],fill=rF))
rc.conditional_formatting.add(f"J3:J{last3}",CellIsRule(operator="greaterThan",formula=["0"],fill=yF))
widths(rc,[9,22,5,14,14,14,14,14,9,11,11,11,14])
rc.freeze_panes="A3"; rc.auto_filter.ref=f"A2:M{last3}"

# ============ BACK_TO_BACK ============
bb=wb.create_sheet("Back_to_Back"); bb.sheet_view.showGridLines=False
bb["A1"]="백투백 Flow-Down: KT–VNPT 계약 의무가 협력사 계약에 동일 반영(flow-down)되는지 점검. 불일치 = KT 중간 리스크."
bb.merge_cells("A1:E1"); bb["A1"].font=Font(name=FONT,bold=True,size=9,color=NAVY); bb["A1"].alignment=wrap; bb.row_dimensions[1].height=26
H4=["조항 / 의무","KT→VNPT (계약상)","협력사 계약 반영(flow-down)","Gap / Action","Owner"]
hdr(bb,H4,row=2)
items=["지급 마일스톤·검수연동","통화 / 환리스크 부담","MG / Rev Share 협력사 배분","하자·페널티·SLA pass-through","IP / License 체인","계약기간·해지 정합","세금 부담·gross-up·분류","납기·물류·통관 책임"]
for i,it in enumerate(items,3):
    bb.cell(row=i,column=1,value=it).font=cfont; 
    for c in range(1,6):
        cell=bb.cell(row=i,column=c); cell.font=cfont; cell.border=B; cell.alignment=wrap
    bb.cell(row=i,column=3).value=""  # Y/N/Gap filled by user
dd(bb,"C",len(items)+2,["반영","부분","미반영"],start=3)
# RAG on flow-down status
bb.conditional_formatting.add(f"C3:C{len(items)+2}",CellIsRule(operator="equal",formula=['"반영"'],fill=gF))
bb.conditional_formatting.add(f"C3:C{len(items)+2}",CellIsRule(operator="equal",formula=['"부분"'],fill=yF))
bb.conditional_formatting.add(f"C3:C{len(items)+2}",CellIsRule(operator="equal",formula=['"미반영"'],fill=rF))
widths(bb,[26,30,18,30,12]); bb.freeze_panes="A3"

wb.save("/home/claude/pack/KT_VNPT_Quotation_Settlement.xlsx")
print("BUILT")
