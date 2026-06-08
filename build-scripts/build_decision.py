from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT="Malgun Gothic"; NAVY="1F4E79"; AMB="FCE4D6"; LT="DDEBF7"; GREY="EDEDED"
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
hfill=PatternFill("solid",fgColor=NAVY); hfont=Font(name=FONT,bold=True,color="FFFFFF",size=10)
cf=Font(name=FONT,size=10); wrap=Alignment(vertical="top",wrap_text=True)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
aF=PatternFill("solid",fgColor=AMB); bF=PatternFill("solid",fgColor=LT); kF=PatternFill("solid",fgColor=GREY)

def thdr(ws,H,row=1):
    for c,h in enumerate(H,1):
        x=ws.cell(row=row,column=c,value=h); x.fill=hfill; x.font=hfont; x.border=B; x.alignment=ctr
    ws.row_dimensions[row].height=30
def widths(ws,w):
    for i,x in enumerate(w,1): ws.column_dimensions[get_column_letter(i)].width=x
def dv(ws,col,n,opts,start=2):
    d=DataValidation(type="list",formula1='"'+",".join(opts)+'"',allow_blank=True); ws.add_data_validation(d); d.add(f"{col}{start}:{col}{n}")

wb=Workbook()

# ---- Guide ----
g=wb.active; g.title="Guide"; g.sheet_view.showGridLines=False
g["B2"]="KT × VNPT — WS1 Working Session Decision Pack"; g["B2"].font=Font(name=FONT,bold=True,size=15,color=NAVY)
g["B3"]="Logistics · After-Sales · SLA  |  RFI #6 후속  |  CONFIDENTIAL"; g["B3"].font=Font(name=FONT,italic=True,size=9,color="808080")
for i,(k,v) in enumerate([
 ("목적","Annex C(RFI #6 답변)의 미확정 [ ] 값을 세션에서 합의하고, OEM에서 받아야 할 데이터를 요청·추적한다."),
 ("탭","① Decision Table(세션 합의용 결정표)  ② OEM Data Request(OEM 입력 요청서)"),
 ("사용법","① Decision Table을 세션 안건으로 사용 → 합의 시 '결정' 입력·Status=Agreed → SLA/계약/견적에 반영.  ② OEM 응답을 받아 AFR·Epidemic·보증·원가 [ ] 채움."),
 ("원칙","정량 SLA·보증·통관 수치는 OEM 백투백·법무(L5) 확인 전까지 placeholder. AS는 3-tier, 지급은 검수연동."),
 ("주의","본 문서의 옵션·범위는 KT 검토용 예시이며 확약이 아니다. 실제 값은 OEM·통관사·변호사 확인 후 확정."),
],start=5):
    g.cell(row=i,column=2,value=k).font=Font(name=FONT,bold=True,size=10,color=NAVY)
    c=g.cell(row=i,column=3,value=v); c.font=cf; c.alignment=wrap
widths(g,[2,12,108])

# ---- Decision Table ----
d=wb.create_sheet("Decision Table"); d.sheet_view.showGridLines=False
H=["#","Item (결정 대상)","Category","Options / Range (KT 검토)","KT 권고","결정 (합의)","Decider","Lands in","Status"]
thdr(d,H)
rows=[
 ("Incoterms","Logistics","CIF / DAP (Hai Phong·Hanoi)","DAP Hanoi","공동","공급/물류계약"),
 ("Importer of Record (IOR)","Logistics","VNPT 직접 / VNPT 지정 디스트리뷰터","VNPT 지정 디스트리뷰터","VNPT","공급계약"),
 ("HS code 분류","Logistics","8528(STB) / 8518(사운드바)","통관사 사전확인","공동","견적·통관"),
 ("FTA·원산지증명(CO)","Logistics","Form VK/AK 적용 여부","적용(관세 경감)","KT","견적·통관"),
 ("안전재고","Logistics","[ ]주 수요 + AFR 예비분","4~6주 + AFR 예비","공동","공급계약"),
 ("보증 기간","Quality-SLA","12 / 18 / 24개월(개통 기준)","STB 24·사운드바 24","공동→OEM","SLA/보증계약"),
 ("DOA 선교환 기한","Quality-SLA","[ ]일","7~14일","KT","SLA"),
 ("AFR 목표 — 8K STB","Quality-SLA","낮은 한 자릿수(OEM 데이터)","OEM 적격치 ≤ X%","공동→OEM","SLA"),
 ("AFR 목표 — 사운드바","Quality-SLA","STB보다 낮음","OEM 적격치 ≤ Y%","공동→OEM","SLA"),
 ("AFR 측정 창","Quality-SLA","12개월 연환산 등","12개월 rolling","공동","SLA"),
 ("Epidemic 임계 N%","Quality-SLA","배포대수의 [N]%","2~5% (협의)","공동→OEM","SLA"),
 ("Epidemic 배수 k×AFR","Quality-SLA","합의 AFR의 [k]배","2~3×","공동→OEM","SLA"),
 ("Epidemic rolling 창","Quality-SLA","[ ]개월","3~6개월","공동","SLA"),
 ("8D RCA 기한","Quality-SLA","트리거 후 [ ]일","5~10영업일","KT/OEM","SLA"),
 ("Epidemic 비용 배분","Quality-SLA","OEM/KT/VNPT 분담·cap","OEM 백투백 우선","공동→OEM/L5","SLA/계약"),
 ("MIC 인증 신청·보유 주체","Certification","KT 스폰서 / 현지 보유자","현지 보유자+KT 기술스폰서","공동","인증 로드맵"),
 ("인증 샘플 수량","Certification","[ ]대","시험기관 요건 확인","KT","인증 로드맵"),
 ("인증 일정·파일럿 게이트","Certification","상용 전 필수 게이트","2026 Q4~ / 파일럿 2027 Q1","공동","일정·게이트"),
 ("예비부품 초기 키트","Logistics","[ ]% of 배포 + AFR 보충","AFR 기반 산정","KT/OEM","공급계약"),
 ("역물류 비용 주체","Logistics","KT / VNPT / 분담","Incoterms 연계 합의","공동","공급계약"),
]
for i,(item,cat,opt,rec,dec,lands) in enumerate(rows,2):
    vals=[i-1,item,cat,opt,rec,"",dec,lands,"Open"]
    for c,v in enumerate(vals,1):
        x=d.cell(row=i,column=c,value=v); x.font=cf; x.border=B; x.alignment=wrap
        if c==1: x.alignment=ctr
        if c==6: x.fill=aF      # 결정 (합의) - to fill
        if c==9: x.alignment=ctr
last=len(rows)+1
dv(d,"C",last,["Logistics","Certification","Quality-SLA"])
dv(d,"G",last,["KT","VNPT","공동","공동→OEM","KT/OEM","공동→OEM/L5"])
dv(d,"I",last,["Open","In discussion","Agreed","Deferred"])
widths(d,[4,24,12,30,22,20,14,14,11]); d.freeze_panes="A2"; d.auto_filter.ref=f"A1:I{last}"

# ---- OEM Data Request ----
o=wb.create_sheet("OEM Data Request"); o.sheet_view.showGridLines=False
H2=["#","Data point (OEM 제공)","SKU / Scope","Needed for (결정)","Unit","OEM 응답","Due","Owner"]
thdr(o,H2)
orows=[
 ("적격 AFR (필드/시험)","8K STB","AFR 목표 SLA","%"),
 ("적격 AFR (필드/시험)","사운드바","AFR 목표 SLA","%"),
 ("신뢰성 시험 데이터(MTBF/HALT)","전 SKU","AFR·보증 근거","-"),
 ("OEM→KT 보증 조건(기간·범위)","전 SKU","보증 기간·백투백","개월"),
 ("Epidemic/대량결함 책임 수용 범위","전 SKU","Epidemic 비용배분","-"),
 ("리콜 비용 분담 조건","전 SKU","Epidemic 비용배분","-"),
 ("단가 (협력사→KT)","STB·사운드바","견적 3단·마진","USD"),
 ("리드타임 / MOQ","전 SKU","안전재고·공급","주/대"),
 ("예비부품 가격·가용성","주요 부품","예비부품 키트","USD"),
 ("수리 TAT (L3 depot)","전 SKU","AS 워크플로우","일"),
 ("국제 시험성적서(CE/FCC 등)","전 SKU","MIC 인증 기초","-"),
]
for i,(dp,scope,need,unit) in enumerate(orows,2):
    vals=[i-1,dp,scope,need,unit,"","",""]
    for c,v in enumerate(vals,1):
        x=o.cell(row=i,column=c,value=v); x.font=cf; x.border=B; x.alignment=wrap
        if c==1: x.alignment=ctr
        if c in (6,7,8): x.fill=aF   # to fill
last2=len(orows)+1
widths(o,[4,28,14,22,8,18,12,12]); o.freeze_panes="A2"; o.auto_filter.ref=f"A1:H{last2}"

wb.save("/home/claude/pack/KT_VNPT_WS1_WorkingSession_Decision_Pack.xlsx")
print("BUILT rows:",len(rows),"oem:",len(orows))
