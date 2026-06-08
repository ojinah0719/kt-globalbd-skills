from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
FONT="Malgun Gothic"; NAVY="1F4E79"; AMBER="FCE4D6"; LTBLUE="DDEBF7"; GREY="EDEDED"
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
hfill=PatternFill("solid",fgColor=NAVY); hfont=Font(name=FONT,bold=True,color="FFFFFF",size=10)
cfont=Font(name=FONT,size=10); wrap=Alignment(vertical="top",wrap_text=True)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
aF=PatternFill("solid",fgColor=AMBER); bF=PatternFill("solid",fgColor=LTBLUE); kF=PatternFill("solid",fgColor=GREY)
MONEY='#,##0.00'; PCT='0.0%'
def info_block(ws,lines,start=4):
    r=start
    for label,val in lines:
        ws.cell(row=r,column=1,value=label).font=Font(name=FONT,bold=True,size=10,color=NAVY)
        ws.cell(row=r,column=2,value=val).font=cfont; r+=1
    return r
def thdr(ws,headers,row):
    for c,h in enumerate(headers,1):
        cell=ws.cell(row=row,column=c,value=h); cell.fill=hfill; cell.font=hfont; cell.border=B; cell.alignment=ctr
    ws.row_dimensions[row].height=30
def widths(ws,ws_w):
    for i,w in enumerate(ws_w,1): ws.column_dimensions[get_column_letter(i)].width=w

# ===== VNPT QUOTATION (WS-based) =====
wb=Workbook(); v=wb.active; v.title="Quotation to VNPT"; v.sheet_view.showGridLines=False
v["A1"]="QUOTATION — KT → VNPT (MyTV)  ·  WS-based draft"; v["A1"].font=Font(name=FONT,bold=True,size=14,color=NAVY)
v["A2"]="KT–VNPT Confidential · Figures ILLUSTRATIVE · WS1=현 단계(Quick Win), WS2–8=Phase 2+(별도 견적)"; v["A2"].font=Font(name=FONT,italic=True,size=9,color="808080")
info=[("From","KT Corporation (Global BD)"),("To","VNPT / MyTV"),("Quotation No","KT-VNPT-2026-[NN]"),("Date","[YYYY-MM-DD]"),
      ("Validity","[ ] days"),("Currency","USD"),("Payment","[milestone / acceptance-linked]"),
      ("FCT note","VNPT withholds FCT(VAT+CIT); per-line classification; Korea–VN DTA may apply"),("Scope","WS1 우선, WS2–8 단계 확장")]
hr=info_block(v,info,4)+1
H=["Line ID","WS","Solution Part","Item","Type","Unit","Qty","Ccy","Unit Price","Line Total","FCT 분류","Notes"]
thdr(v,H,hr)
lines=[
 ("V-001","WS1","STB","Genie STB 4 단말 (OEM)","goods","ea",1000,"USD",110,"goods (1%+1%)","ILLUSTRATIVE"),
 ("V-002","WS1","Soundbar","All-in-one Soundbar 8K","goods","ea",500,"USD",60,"goods (1%+1%)","ILLUSTRATIVE"),
 ("V-003","WS1","Media AI Agent","AI Agent License","royalty","license",1,"USD",57500,"royalty CIT 10%; DTA","ILLUSTRATIVE"),
 ("V-004","WS1","TMS","Terminal Mgmt System","royalty","license",1,"USD",30000,"royalty CIT 10%","ILLUSTRATIVE"),
 ("V-005","WS1","Launcher/OS","Custom launcher / GTV 연동","service","LS",1,"USD",20000,"service 5%+5%","ILLUSTRATIVE"),
 ("V-006","WS1","SI/Integration","연동·구축","service","LS",1,"USD",89600,"service 5%+5%","ILLUSTRATIVE"),
 ("V-007","WS1","Localization","현지화","service","LS",1,"USD",25000,"service 5%+5%","ILLUSTRATIVE"),
 ("V-008","WS1","Certification","MIC ICT + Google 인증","service","LS",1,"USD",15000,"service","ILLUSTRATIVE"),
 ("V-101","WS2","Recommendation","추천 엔진","royalty","license",1,"USD",None,"royalty","Phase 2+ (별도 견적)"),
 ("V-102","WS3","AI Content Ops","콘텐츠 운영 AI","service","LS",1,"USD",None,"service","Phase 2+"),
 ("V-103","WS4","K-Content","K콘텐츠 소싱","royalty","license",1,"USD",None,"royalty/license","Phase 2+"),
 ("V-104","WS5","Advertising","광고 플랫폼","service","LS",1,"USD",None,"Rev Share","Phase 2+"),
 ("V-105","WS6","OTT Aggregation","OTT 번들/Aggregation","royalty","LS",1,"USD",None,"royalty/service","Phase 2+"),
 ("V-106","WS7","HiOrder","호스피탈리티(테이블오더)","goods","set",1,"USD",None,"goods+royalty","Phase 2+"),
 ("V-107","WS8","B2G/Learning","교육·B2G","service","LS",1,"USD",None,"service","Phase 2+"),
]
first=hr+1
for i,row in enumerate(lines,first):
    fill=aF if row[1]=="WS1" else kF
    for c,val in enumerate(row,1):
        cell=v.cell(row=i,column=c,value=val); cell.font=cfont; cell.border=B; cell.alignment=wrap; cell.fill=fill
last=first+len(lines)-1
for i in range(first,last+1):
    v.cell(row=i,column=10,value=f"=IFERROR(I{i}*G{i},\"\")"); v.cell(row=i,column=9).number_format=MONEY; v.cell(row=i,column=10).number_format=MONEY
v.cell(row=last+1,column=3,value="TOTAL WS1 (USD)").font=Font(name=FONT,bold=True)
t=v.cell(row=last+1,column=10,value=f"=SUM(J{first}:J{first+7})"); t.font=Font(name=FONT,bold=True); t.number_format=MONEY; t.fill=bF; t.border=B
v.cell(row=last+1,column=3).fill=bF; v.cell(row=last+1,column=3).border=B
tr=last+3
v.cell(row=tr,column=1,value="Terms / Notes").font=Font(name=FONT,bold=True,size=10,color=NAVY)
for j,txt in enumerate(["· KT→VNPT 견적 — 협력사 원가·KT 내부 마진 미표기.",
  "· WS1(Device&AI UX)=현 단계 Quick Win. WS2–8은 WS1 기반 단계 확장(별도 견적).",
  "· 세금: VNPT가 FCT(VAT+CIT) 원천징수, 라인별 분류(goods/royalty/service), 한–베 DTA 검토 — 세무사 확인.",
  "· 데이터 PDPL 2025·Decree 356(VNPT Cloud/DPIA/CTIA), 인증 MIC ICT·Google.",
  "· 금액 ILLUSTRATIVE — 정식 견적·세무·계약 별도(법무: L5-bd-contract-review)."],1):
    v.cell(row=tr+j,column=1,value=txt).font=cfont
widths(v,[9,6,16,22,9,6,7,6,13,14,20,18]); v.freeze_panes=f"A{first}"
wb.save("/home/claude/pack/KT_Quotation_to_VNPT.xlsx"); print("VNPT BUILT")

# ===== PARTNER RFQ (WS-based, with Partner col) =====
wb2=Workbook(); ws=wb2.active; ws.title="RFQ to Partner"; ws.sheet_view.showGridLines=False
ws["A1"]="QUOTATION (RFQ) — KT ↔ Partners  ·  WS-based draft"; ws["A1"].font=Font(name=FONT,bold=True,size=14,color=NAVY)
ws["A2"]="KT–Partner Confidential · Figures ILLUSTRATIVE · VNPT가격·KT마진 미표기 · 파트별 협력사 단가"; ws["A2"].font=Font(name=FONT,italic=True,size=9,color="808080")
info2=[("From","KT Corporation (Global BD)"),("To","[각 협력사]"),("RFQ No","RFQ-2026-[NN]"),("Date","[YYYY-MM-DD]"),
       ("Validity","[ ]일"),("Currency","KRW / USD"),("Incoterms(HW)","[예: DAP Hanoi]"),("Payment","[검수 후 NN일]")]
hr2=info_block(ws,info2,4)+1
H2=["Line ID","WS","Solution Part","Item / Spec","Partner","Type","Unit","Qty","Ccy","단가","Line Total","VAT%","Total incl VAT","Notes"]
thdr(ws,H2,hr2)
plines=[
 ("P-001","WS1","STB","Genie STB 4","[STB OEM]","goods","ea",1000,"USD",100,0.10,"ILLUSTRATIVE"),
 ("P-002","WS1","Soundbar","Soundbar 8K","[Audio OEM]","goods","ea",500,"USD",52,0.10,"ILLUSTRATIVE"),
 ("P-003","WS1","Media AI Agent","AI Agent SW","[AI vendor / KT]","royalty","license",1,"USD",50000,0.0,"ILLUSTRATIVE"),
 ("P-004","WS1","TMS","TMS SW","[KT / vendor]","royalty","license",1,"USD",26000,0.0,"ILLUSTRATIVE"),
 ("P-005","WS1","Launcher/OS","Launcher 개발","[SW house]","service","LS",1,"USD",17000,0.10,"ILLUSTRATIVE"),
 ("P-006","WS1","SI/Integration","구축","[SI partner]","service","LS",1,"USD",78000,0.10,"ILLUSTRATIVE"),
 ("P-007","WS1","Localization","현지화","[Local vendor]","service","LS",1,"USD",21000,0.10,"ILLUSTRATIVE"),
 ("P-008","WS1","Certification","인증 지원","[Test lab]","service","LS",1,"USD",12000,0.10,"ILLUSTRATIVE"),
]
f2=hr2+1
for i,row in enumerate(plines,f2):
    # row: id,ws,part,item,partner,type,unit,qty,ccy,price,vat,notes  -> cols A..L then K(total),M(incl) formulas
    vals=[row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],None,row[10],None,row[11]]
    for c,val in enumerate(vals,1):
        cell=ws.cell(row=i,column=c,value=val); cell.font=cfont; cell.border=B; cell.alignment=wrap; cell.fill=aF
l2=f2+15
for i in range(f2,l2+1):
    ws.cell(row=i,column=11,value=f"=IFERROR(J{i}*H{i},\"\")")          # Line Total = 단가*Qty
    ws.cell(row=i,column=13,value=f"=IFERROR(K{i}*(1+L{i}),\"\")")      # incl VAT
    for cc in (10,11,13): ws.cell(row=i,column=cc).number_format=MONEY
    ws.cell(row=i,column=12).number_format=PCT
ws.cell(row=l2+1,column=3,value="TOTAL").font=Font(name=FONT,bold=True)
for col in (11,13):
    L=get_column_letter(col); t=ws.cell(row=l2+1,column=col,value=f"=SUM({L}{f2}:{L}{l2})"); t.font=Font(name=FONT,bold=True); t.number_format=MONEY; t.fill=bF; t.border=B
ws.cell(row=l2+1,column=3).fill=bF; ws.cell(row=l2+1,column=3).border=B
tr2=l2+3
ws.cell(row=tr2,column=1,value="조건/비고").font=Font(name=FONT,bold=True,size=10,color=NAVY)
for j,txt in enumerate(["· 협력사↔KT 견적(WS1 파트별). VNPT 가격·KT 마진 미포함.",
  "· 세금: 한국 부가세 기준(해외 협력사는 원천·이전가격 별도). 검수 연동 지급.",
  "· 파트별 협력사는 [ ]에 지정. 통화·환율 기준일·유효기간 명시.",
  "· 금액 ILLUSTRATIVE — 정식 견적·세무 별도."],1):
    ws.cell(row=tr2+j,column=1,value=txt).font=cfont
widths(ws,[9,6,15,18,15,8,6,7,6,12,13,7,14,14]); ws.freeze_panes=f"A{f2}"
wb2.save("/home/claude/pack/KT_Quotation_to_Partner_RFQ.xlsx"); print("PARTNER BUILT")
