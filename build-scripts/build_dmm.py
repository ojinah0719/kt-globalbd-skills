from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT = "Malgun Gothic"; BLUE = "1F4E79"
GREEN, YELLOW, RED, GREY, AMBER, LTBLUE = "D5E8D4", "FFF2CC", "F8CECC", "EDEDED", "FCE4D6", "DDEBF7"
thin = Side(style="thin", color="BFBFBF"); B = Border(left=thin, right=thin, top=thin, bottom=thin)
hfill = PatternFill("solid", fgColor=BLUE); hfont = Font(name=FONT, bold=True, color="FFFFFF", size=10)
cfont = Font(name=FONT, size=10); wrap = Alignment(vertical="top", wrap_text=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
gF, yF, rF, kF, aF, bF = (PatternFill("solid", fgColor=x) for x in (GREEN, YELLOW, RED, GREY, AMBER, LTBLUE))

wb = Workbook()

def hdrow(ws, ncol, row=1):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c); cell.fill = hfill; cell.font = hfont; cell.border = B; cell.alignment = center
    ws.row_dimensions[row].height = 30

def rows_in(ws, data, start=2):
    for r, rv in enumerate(data, start=start):
        for c, v in enumerate(rv, start=1):
            cell = ws.cell(row=r, column=c, value=v); cell.font = cfont; cell.border = B; cell.alignment = wrap

def widths(ws, ws_w):
    for i, w in enumerate(ws_w, start=1): ws.column_dimensions[get_column_letter(i)].width = w

def rag(ws, col, n, green=(), yellow=(), red=(), grey=(), amber=(), blue=()):
    rng = f"{col}2:{col}{n+1}"
    for vals, f in [(green, gF), (yellow, yF), (red, rF), (grey, kF), (amber, aF), (blue, bF)]:
        for v in vals: ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{v}"'], fill=f))

def dd(ws, col, n, opts):
    dv = DataValidation(type="list", formula1='"' + ",".join(opts) + '"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"{col}2:{col}{n+1}")

# ============ REGISTER (SSOT) ============
reg = wb.active; reg.title = "Register"
hdr = ["DEL-ID","Deliverable name","Stage","Phase","Type","Author (R)","Approver (A)","Version","Status","Confidentiality","Reg. date","Approved date","Storage location (link)","Linked RTM","Linked DEC","Supersedes","Notes"]
reg.append(hdr); hdrow(reg, len(hdr))
L = "[link]"
data = [
 ["DEL-001","KT Media Solution Proposal","1","P0","Proposal","KT","KT","v1.0","Baselined","Confidential","2026-03-17","2026-03-17","/01_Proposal/"+L,"","","","Initial solution proposal"],
 ["DEL-002","MyTV Collaborative Discussion (received)","2","P0","Input","VNPT/MyTV","","v1.0","Baselined","Confidential","2026-03-26","","/01_Proposal/"+L,"","","","VNPT needs (8 WS) + role split"],
 ["DEL-003","Business Cooperation Proposal","2","P0","Proposal","KT","KT","v1.0","Baselined","Confidential","2026-04-30","2026-04-30","/01_Proposal/"+L,"","","","8-WS analysis & strategy"],
 ["DEL-004","VNPT RFI (WS01) (received)","2","P0","Input","VNPT/MyTV","","v1.0","Baselined","Confidential","2026-05-15","","/02_Consulting/"+L,"","","","7-item RFI"],
 ["DEL-005","WS1 Joint Working Session Package","3","P0","Plan","KT (PMO)","KT","v1.0","Approved","Confidential","2026-06-06","2026-06-06","/02_Consulting/"+L,"RTM-01..07","","","Meeting package"],
 ["DEL-006","KT Briefing Deck (BM/STB/AI Agent·TMS)","3","P0","Presentation","KT","","v0.1","Drafting","Confidential","2026-06-06","","/02_Consulting/"+L,"RTM-01..05","DEC-01..03","","Placeholders for pricing/MG"],
 ["DEL-007","PMO Trackers workbook","3","P0","Tracker","KT (PMO)","KT","v1.0","Approved","Confidential","2026-06-06","2026-06-06","/00_PMO/"+L,"","","","RTM/RAID/Decision/Registry"],
 ["DEL-008","RFI 7-item response (RTM)","3","P0","Matrix","KT","Joint","v0.3","Drafting","Confidential","2026-06-06","","/02_Consulting/"+L,"RTM-01..07","","","Working draft"],
 ["DEL-009","STB pricing & MG model","3","P0","Model","KT","KT","v0.1","Drafting","Restricted","","","/02_Consulting/"+L,"RTM-04","DEC-01","","Commercial-sensitive"],
 ["DEL-010","STB deployment plan","3","P1","Plan","KT","Joint","v0.1","Drafting","Confidential","","","/02_Consulting/"+L,"RTM-01,03","AI-03","","-"],
 ["DEL-011","Media AI Agent & TMS scope/impl. plan","3","P1","Spec","KT","Joint","v0.1","Drafting","Confidential","","","/02_Consulting/"+L,"RTM-02,05","","","-"],
 ["DEL-012","Mutual NDA (VNPT standard terms)","3","P0","Legal","VNPT/MyTV","Both · Legal","\u2014","In review","Restricted","","","/05_Legal/"+L,"","DEC-05","","VNPT standard terms"],
 ["DEL-013","Requirements analysis / SRS","4-5","P1","Spec","KT","Joint","\u2014","Drafting","Confidential","","","/03_Requirements/"+L,"RTM-01..07","","","Planned"],
 ["DEL-014","Scope definition (SOW)","5","P1","Plan","Joint","Joint","\u2014","Drafting","Confidential","","","/03_Requirements/"+L,"","","","Contract annex; planned"],
 ["DEL-015","Detailed service scenario","6","P1","Spec","MyTV","Joint","\u2014","Drafting","Confidential","","","/04_Design/"+L,"","","","Planned"],
 ["DEL-016","Interface Definition Document (IDD) / System Design","7","P1","Design","KT","Joint","\u2014","Drafting","Confidential","","","/04_Design/"+L,"RTM-02,05","","","Planned"],
 ["DEL-017","Certification & regulatory roadmap","10","P2","Plan","Joint","Joint","\u2014","Drafting","Confidential","","","/06_Cert/"+L,"RTM-03,06,07","DEC-02","","Planned (MIC/Google)"],
 ["DEL-018","Data sovereignty design","7","P2","Design","Joint","Joint","\u2014","Drafting","Restricted","","","/04_Design/"+L,"RTM-07","DEC-02","","Decree 13/2023"],
 ["DEL-019","Test & UAT results","9","P3","Report","KT","Joint","\u2014","Drafting","Confidential","","","/07_Test/"+L,"","","","Planned"],
 ["DEL-020","MOU","5","P1","Legal","Joint","Both · Legal","\u2014","Drafting","Restricted","","","/05_Legal/"+L,"","DEC-04","","Target Q3 2026"],
]
N = len(data)
rows_in(reg, data)
widths(reg, [9,34,7,7,12,11,12,8,11,14,11,12,22,13,13,11,26])
rag(reg, "I", N, green=["Approved","Baselined"], yellow=["Drafting","In review"], grey=["Superseded"])
rag(reg, "J", N, red=["Restricted"], amber=["Confidential"], blue=["Internal"], grey=["Public"])
dd(reg, "I", N, ["Drafting","In review","Approved","Baselined","Superseded"])
dd(reg, "J", N, ["Public","Internal","Confidential","Restricted"])
dd(reg, "E", N, ["Proposal","Report","Spec","Plan","Matrix","Design","Model","Legal","Minutes","Tracker","Presentation","Input"])
dd(reg, "D", N, ["P0","P1","P2","P3","P4"])
dd(reg, "F", N, ["KT","KT (PMO)","MyTV","VNPT/MyTV","Joint","Both · Legal"])
reg.auto_filter.ref = f"A1:{get_column_letter(len(hdr))}{N+1}"
reg.freeze_panes = "C2"; reg.sheet_view.showGridLines = False

# ============ DASHBOARD (formulas) ============
db = wb.create_sheet("Dashboard", 0)
db.sheet_view.showGridLines = False
db["B2"] = "Document Management Master \u2014 Dashboard"; db["B2"].font = Font(name=FONT, bold=True, size=16, color=BLUE)
db["B3"] = "KT \u00D7 VNPT/MyTV  |  Single Source of Truth for all project deliverables  |  CONFIDENTIAL"
db["B3"].font = Font(name=FONT, italic=True, size=9, color="808080")
db["B5"] = "Total deliverables"; db["B5"].font = Font(name=FONT, bold=True, size=11)
db["D5"] = "=COUNTA(Register!A2:A1000)"; db["D5"].font = Font(name=FONT, bold=True, size=14, color=BLUE)
# status block
db["B7"] = "By status"; db["B7"].font = Font(name=FONT, bold=True, size=11, color=BLUE)
stat = ["Drafting","In review","Approved","Baselined","Superseded"]
for i, sname in enumerate(stat):
    r = 8 + i
    c1 = db.cell(row=r, column=2, value=sname); c1.font = cfont; c1.border = B
    c2 = db.cell(row=r, column=3, value=f'=COUNTIF(Register!I2:I1000,"{sname}")'); c2.font = cfont; c2.border = B; c2.alignment = center
    c1.fill = {"Approved": gF, "Baselined": gF, "Drafting": yF, "In review": yF, "Superseded": kF}[sname]
db["B14"] = "% Approved or Baselined"; db["B14"].font = Font(name=FONT, bold=True, size=11)
db["D14"] = "=(C10+C11)/D5"; db["D14"].number_format = "0%"; db["D14"].font = Font(name=FONT, bold=True, size=14, color="548235")
# phase block
db["F7"] = "By phase"; db["F7"].font = Font(name=FONT, bold=True, size=11, color=BLUE)
for i, ph in enumerate(["P0","P1","P2","P3","P4"]):
    r = 8 + i
    c1 = db.cell(row=r, column=6, value=ph); c1.font = cfont; c1.border = B
    c2 = db.cell(row=r, column=7, value=f'=COUNTIF(Register!D2:D1000,"{ph}")'); c2.font = cfont; c2.border = B; c2.alignment = center
# confidentiality
db["F14"] = "Restricted items"; db["F14"].font = Font(name=FONT, bold=True, size=11)
db["G14"] = '=COUNTIF(Register!J2:J1000,"Restricted")'; db["G14"].font = Font(name=FONT, bold=True, size=12, color="B00020"); db["G14"].alignment = center; db["G14"].border = B
db["B17"] = "How to use"; db["B17"].font = Font(name=FONT, bold=True, size=11, color=BLUE)
note = ("1) Every deliverable = one row in 'Register' (the single source of truth) \u2014 unregistered = not official.   "
        "2) Follow ID / version / file-naming & status rules on the 'Conventions' tab.   "
        "3) Status flows Drafting \u2192 In review \u2192 Approved \u2192 Baselined; approval is gated at each Stage-Gate.   "
        "4) Record every version change in 'Change_Log'.   "
        "5) Review this register at every meeting and gate. Counts above auto-update from 'Register'.")
db["B18"] = note; db["B18"].font = cfont; db["B18"].alignment = wrap
db.merge_cells("B18:G22")
widths(db, [2, 22, 12, 10, 14, 12, 12])

# ============ CHANGE_LOG ============
cl = wb.create_sheet("Change_Log")
clh = ["Date","DEL-ID","Version","Change description","Author","Approved by"]
cl.append(clh); hdrow(cl, len(clh))
cld = [
 ["2026-03-17","DEL-001","v1.0","Initial issue","KT","KT"],
 ["2026-04-30","DEL-003","v1.0","Initial issue","KT","KT"],
 ["2026-06-06","DEL-005","v1.0","Issued for working session","KT (PMO)","KT"],
 ["2026-06-06","DEL-006","v0.1","Draft \u2014 commercial figures as placeholders","KT","\u2014"],
 ["2026-06-06","DEL-007","v1.0","Trackers seeded (RTM/RAID/Decision/Registry)","KT (PMO)","KT"],
 ["2026-06-06","DEL-008","v0.3","RFI response draft","KT","\u2014"],
]
rows_in(cl, cld); widths(cl, [12,10,9,46,14,14])
cl.auto_filter.ref = f"A1:F{len(cld)+1}"; cl.freeze_panes = "A2"; cl.sheet_view.showGridLines = False

# ============ CONVENTIONS ============
cv = wb.create_sheet("Conventions"); cv.sheet_view.showGridLines = False
def head(r, txt): cv.cell(row=r, column=2, value=txt).font = Font(name=FONT, bold=True, size=12, color=BLUE)
def line(r, txt, bold=False, col=2): 
    c = cv.cell(row=r, column=col, value=txt); c.font = Font(name=FONT, size=10, bold=bold); c.alignment = wrap
cv.cell(row=2, column=2, value="Document Management \u2014 Conventions").font = Font(name=FONT, bold=True, size=15, color=BLUE)
head(4, "1. ID rule")
line(5, "DEL-### sequential (e.g., DEL-001). Optionally stage-scoped: S07-DEL-01. Never reuse an ID; superseded docs keep their ID with status 'Superseded'.")
head(7, "2. Version rule")
line(8, "v0.x = draft / in review.   v1.0 = approved & baselined.   Increment minor for edits, major for re-baseline. Record each change in Change_Log.")
head(10, "3. File naming")
line(11, "KTxVNPT_WS1_<DeliverableName>_v<Version>_<YYYYMMDD>_<Confidentiality>.<ext>")
line(12, "e.g.  KTxVNPT_WS1_BriefingDeck_v0.1_20260606_Confidential.pptx", bold=True)
head(14, "4. Status workflow")
line(15, "Drafting  \u2192  In review  \u2192  Approved  \u2192  Baselined  \u2192  (Superseded). Approval is gated at each Stage-Gate: a stage advances only when its deliverables are Approved.")
head(17, "5. Confidentiality grades")
line(18, "Public  /  Internal  /  Confidential (NDA-covered)  /  Restricted (commercial & legal: pricing, MG, contracts, personal data).")
head(20, "6. Folder structure (Teams / SharePoint)")
for i, f in enumerate(["/00_PMO  (trackers, register, governance)","/01_Proposal","/02_Consulting (RFI, working session)","/03_Requirements","/04_Design (IDD, scenarios, data sovereignty)","/05_Legal (NDA, MOU, contract)","/06_Cert (MIC, Google, compliance)","/07_Test","/08_Pilot_Commercial"]):
    line(21 + i, f)
head(31, "7. Roles")
line(32, "Document Controller = PMO (owns the register).   Author (R) creates;   Approver (A) approves at gate.   All members register before sharing.")
head(34, "8. Registration process")
line(35, "Create \u2192 name per rule \u2192 add a Register row (Drafting) \u2192 store in the matching folder & paste link \u2192 review \u2192 Approve/Baseline at gate \u2192 log change \u2192 supersede old version if any.")
head(37, "9. Traceability")
line(38, "Each deliverable links to its requirement (RTM) and decision (Decision Log), which carry through to the contract. The register is reviewed at every meeting and Stage-Gate.")
widths(cv, [2, 120])

wb.save("/home/claude/pack/KT_VNPT_Document_Management_Master.xlsx")
print("DMM BUILT")
